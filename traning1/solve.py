"""Reproducible MILP solution for Training Problem 14.

The model uses actual hourly processing and backlog balance; unused capacity cannot
be carried into future hours. Run this file to regenerate all tables and figures.
"""
from __future__ import annotations

import csv
import json
import math
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook, load_workbook
from scipy.optimize import Bounds, LinearConstraint, milp
from scipy.sparse import lil_matrix

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "generated"
FIG = ROOT / "paper" / "figures"
OUT.mkdir(exist_ok=True)
FIG.mkdir(exist_ok=True)

wb = load_workbook(ROOT / "附件.xlsx", data_only=True)
raw = list(wb.active.iter_rows(min_row=2, values_only=True))
arrivals = np.zeros((30, 24), dtype=float)
for day, hour, amount in raw:
    arrivals[int(day) - 1, int(hour)] = float(amount)

STARTS = tuple(range(17))
DURATION = 8
BIG_M = 1000


def _add_row(rows, lower, upper, terms, lo=-np.inf, hi=np.inf):
    r = len(lower)
    lower.append(lo)
    upper.append(hi)
    for j, value in terms:
        rows[r, j] += value


def solve_day(day: int, degraded: bool, normal_rate: float = 25.0, low_rate: float = 10.0) -> dict:
    """Solve one day; degraded=True adds Q2's one low-productivity hour."""
    a = arrivals[day]
    nx, ny = 17, 17
    nz = 17 * 8 if degraded else 0
    npv, nb = 24, 24
    ox, oy, oz, op, ob = 0, nx, nx + ny, nx + ny + nz, nx + ny + nz + npv
    n = ob + nb

    c = np.zeros(n)
    c[oy:oz] = 1.0
    integrality = np.zeros(n)
    integrality[ox:op] = 1
    lb = np.zeros(n)
    ub = np.full(n, np.inf)
    ub[ox:oy] = 1
    ub[oy:op] = BIG_M

    # Allocate more rows than needed; trim before solve.
    rows = lil_matrix((300, n), dtype=float)
    lower, upper = [], []

    _add_row(rows, lower, upper, [(ox + s, 1) for s in STARTS], 5, 5)
    for s in STARTS:
        _add_row(rows, lower, upper, [(oy + s, 1), (ox + s, -BIG_M)], hi=0)
        _add_row(rows, lower, upper, [(oy + s, 1), (ox + s, -1)], lo=0)
        if degraded:
            _add_row(
                rows, lower, upper,
                [(oz + 8 * s + k, 1) for k in range(8)] + [(oy + s, -1)],
                0, 0,
            )

    penalty = normal_rate - low_rate
    # Actual processing cannot exceed capacity in that hour.
    for h in range(24):
        terms = [(op + h, 1)]
        for s in STARTS:
            if s <= h < s + DURATION:
                terms.append((oy + s, -normal_rate))
                if degraded:
                    terms.append((oz + 8 * s + (h - s), penalty))
        _add_row(rows, lower, upper, terms, hi=0)

    # Backlog flow: b_h = b_(h-1) + a_h - p_h, b_h >= 0, b_23 = 0.
    for h in range(24):
        terms = [(ob + h, 1), (op + h, 1)]
        if h > 0:
            terms.append((ob + h - 1, -1))
        _add_row(rows, lower, upper, terms, a[h], a[h])
    _add_row(rows, lower, upper, [(ob + 23, 1)], 0, 0)

    if degraded:
        # By 16:00, processed volume must cover all arrivals from hours 0--12.
        _add_row(
            rows, lower, upper,
            [(op + h, 1) for h in range(16)],
            float(a[:13].sum()), np.inf,
        )

    m = len(lower)
    result = milp(
        c,
        integrality=integrality,
        bounds=Bounds(lb, ub),
        constraints=LinearConstraint(rows[:m].tocsr(), np.array(lower), np.array(upper)),
        options={"time_limit": 60, "mip_rel_gap": 0},
    )
    if not result.success:
        raise RuntimeError(f"day {day + 1}: {result.message}")

    v = result.x
    x = np.rint(v[ox:oy]).astype(int)
    y = np.rint(v[oy:oz]).astype(int)
    p = v[op:ob]
    b = v[ob:ob + 24]
    z = np.rint(v[oz:op]).astype(int).reshape(17, 8) if degraded else np.zeros((17, 8), int)

    # Independent post-solve checks.
    capacity = np.zeros(24)
    for h in range(24):
        for s in STARTS:
            if s <= h < s + 8:
                capacity[h] += normal_rate * y[s]
                if degraded:
                    capacity[h] -= penalty * z[s, h - s]
    assert x.sum() == 5 and np.all(y <= BIG_M * x) and np.all(y >= x)
    assert np.max(p - capacity) <= 1e-5
    assert np.min(b) >= -1e-5 and abs(b[-1]) <= 1e-5
    assert np.max(np.abs(b - (np.cumsum(a) - np.cumsum(p)))) <= 1e-4
    if degraded:
        assert np.all(z.sum(axis=1) == y)
        assert p[:16].sum() + 1e-5 >= a[:13].sum()

    return {
        "day": day + 1,
        "objective": int(y.sum()),
        "selected": [s for s in STARTS if x[s]],
        "workers": {str(s): int(y[s]) for s in STARTS if x[s]},
        "low_productivity": {
            str(s): {str(k): int(z[s, k]) for k in range(8) if z[s, k]}
            for s in STARTS if x[s]
        } if degraded else {},
        "processing": [round(float(q), 4) for q in p],
        "backlog": [round(max(0.0, float(q)), 4) for q in b],
        "capacity": [round(float(q), 4) for q in capacity],
    }


def solve_monthly_roster(requirements: list[int]) -> tuple[int, np.ndarray]:
    """Minimum headcount and a concrete 0/1 work roster for Q3."""
    lower_bound = max(max(requirements), math.ceil(sum(requirements) / 23))
    for n_workers in range(lower_bound, lower_bound + 20):
        nvar = n_workers * 30
        c = np.zeros(nvar)
        integ = np.ones(nvar)
        lb = np.zeros(nvar)
        ub = np.ones(nvar)
        nrows = n_workers + n_workers * 23 + 30
        A = lil_matrix((nrows, nvar), dtype=float)
        lo, hi = [], []

        def idx(i, d): return i * 30 + d

        for i in range(n_workers):
            _add_row(A, lo, hi, [(idx(i, d), 1) for d in range(30)], 23, 23)
            for start in range(23):
                _add_row(A, lo, hi, [(idx(i, d), 1) for d in range(start, start + 8)], hi=7)
        for d, need in enumerate(requirements):
            _add_row(A, lo, hi, [(idx(i, d), 1) for i in range(n_workers)], need, np.inf)

        result = milp(
            c,
            integrality=integ,
            bounds=Bounds(lb, ub),
            constraints=LinearConstraint(A.tocsr(), np.array(lo), np.array(hi)),
            options={"time_limit": 180, "mip_rel_gap": 0},
        )
        if result.success:
            roster = np.rint(result.x).astype(int).reshape(n_workers, 30)
            assert np.all(roster.sum(axis=1) == 23)
            for i in range(n_workers):
                assert max(roster[i, s:s + 8].sum() for s in range(23)) <= 7
            assert np.all(roster.sum(axis=0) >= np.array(requirements))
            return n_workers, roster
    raise RuntimeError("monthly roster was not feasible in searched range")


def assign_shifts(q2: list[dict], roster: np.ndarray) -> list[dict]:
    """Partition each day's working employees among the five Q2 shifts."""
    allocation = []
    for d, solution in enumerate(q2):
        on_duty = (np.flatnonzero(roster[:, d]) + 1).tolist()
        required = dict(solution["workers"])
        extra = len(on_duty) - sum(required.values())
        largest = max(required, key=required.get)
        required[largest] += extra
        cursor = 0
        day_alloc = {}
        for s in sorted(required, key=int):
            count = required[s]
            day_alloc[s] = on_duty[cursor:cursor + count]
            cursor += count
        assert cursor == len(on_duty)
        allocation.append({"day": d + 1, "shifts": day_alloc})
    return allocation


def save_outputs(q1, q2, n_workers, roster, allocation, sensitivity):
    result = {"q1": q1, "q2": q2, "q3": {"minimum_workers": n_workers}, "sensitivity": sensitivity}
    (OUT / "results.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    with (OUT / "daily_summary.csv").open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["day", "arrivals", "q1_workers", "q2_workers", "q1_starts", "q2_starts"])
        for d in range(30):
            w.writerow([d + 1, int(arrivals[d].sum()), q1[d]["objective"], q2[d]["objective"],
                        "/".join(map(str, q1[d]["selected"])), "/".join(map(str, q2[d]["selected"]))])

    book = Workbook()
    ws = book.active
    ws.title = "每日汇总"
    ws.append(["日期", "进货总量", "问题一人数", "问题二人数", "问题一班次起点", "问题二班次起点"])
    for d in range(30):
        ws.append([d + 1, int(arrivals[d].sum()), q1[d]["objective"], q2[d]["objective"],
                   ",".join(map(str, q1[d]["selected"])), ",".join(map(str, q2[d]["selected"]))])
    wr = book.create_sheet("问题三出勤表")
    wr.append(["工人编号"] + [f"第{d}天" for d in range(1, 31)] + ["工作天数"])
    for i, row in enumerate(roster, 1):
        wr.append([i] + row.tolist() + [int(row.sum())])
    wa = book.create_sheet("问题三班次分配")
    wa.append(["日期", "班次起点", "人数", "工人编号"])
    for item in allocation:
        for s, ids in item["shifts"].items():
            wa.append([item["day"], int(s), len(ids), ",".join(map(str, ids))])
    book.save(OUT / "排班结果.xlsx")

    # LaTeX tables consumed directly by the paper.
    table_lines = [
        r"\begin{longtable}{rrrrrr}",
        r"\caption{30天逐日最优排班汇总}\label{tab:daily}\\",
        r"\toprule 日期 & 到货量 & 问题一人数 & 问题二人数 & 问题一班次起点 & 问题二班次起点\\\midrule",
        r"\endfirsthead\toprule 日期 & 到货量 & 问题一人数 & 问题二人数 & 问题一班次起点 & 问题二班次起点\\\midrule\endhead",
    ]
    for d in range(30):
        table_lines.append(
            f"{d+1} & {int(arrivals[d].sum())} & {q1[d]['objective']} & {q2[d]['objective']} & "
            f"{'/'.join(map(str,q1[d]['selected']))} & {'/'.join(map(str,q2[d]['selected']))}\\\\"
        )
    table_lines += [r"\bottomrule", r"\end{longtable}"]
    (OUT / "daily_table.tex").write_text("\n".join(table_lines), encoding="utf-8")

    scheduled = roster.sum(axis=0)
    roster_lines = [
        r"\begin{table}[htbp]\centering",
        r"\caption{问题三月度排班可行性检验}\label{tab:roster-check}",
        r"\begin{tabular}{lrr}\toprule 指标 & 要求 & 计算结果\\\midrule",
        f"招工人数 & 最少 & {n_workers}\\\\",
        f"每人工日数 & 23 & {int(roster.sum(axis=1).min())}--{int(roster.sum(axis=1).max())}\\\\",
        f"连续工作上限 & 不超过7天 & {max(max(roster[i,s:s+8].sum() for s in range(23)) for i in range(n_workers))}\\\\",
        f"全月总工日 & 不低于需求 & {int(scheduled.sum())}（需求{sum(x['objective'] for x in q2)}）\\\\",
        r"\bottomrule\end{tabular}\end{table}",
    ]
    (OUT / "roster_check.tex").write_text("\n".join(roster_lines), encoding="utf-8")

    # Figures.
    days = np.arange(1, 31)
    fig, ax1 = plt.subplots(figsize=(9, 4.6))
    ax1.bar(days, arrivals.sum(axis=1), color="#d9e6f2", label="Daily arrivals")
    ax1.set_xlabel("Day"); ax1.set_ylabel("Arrivals")
    ax2 = ax1.twinx()
    ax2.plot(days, [x["objective"] for x in q1], "o-", lw=1.5, ms=3, label="Q1 workers", color="#2166ac")
    ax2.plot(days, [x["objective"] for x in q2], "s-", lw=1.5, ms=3, label="Q2 workers", color="#b2182b")
    ax2.set_ylabel("Workers")
    lines, labels = ax1.get_legend_handles_labels(); lines2, labels2 = ax2.get_legend_handles_labels()
    ax2.legend(lines + lines2, labels + labels2, loc="upper right", frameon=False)
    fig.tight_layout(); fig.savefig(FIG / "daily_demand_workers.pdf", bbox_inches="tight"); plt.close(fig)

    fig, ax = plt.subplots(figsize=(9, 5))
    im = ax.imshow(arrivals, aspect="auto", cmap="YlOrRd")
    ax.set_xlabel("Hour"); ax.set_ylabel("Day")
    ax.set_xticks(range(0, 24, 2)); ax.set_xticklabels(range(0, 24, 2))
    ax.set_yticks(range(0, 30, 3)); ax.set_yticklabels(range(1, 31, 3))
    fig.colorbar(im, ax=ax, label="Hourly arrivals")
    fig.tight_layout(); fig.savefig(FIG / "arrival_heatmap.pdf", bbox_inches="tight"); plt.close(fig)

    fig, ax = plt.subplots(figsize=(9, 4.2))
    ax.plot(days, roster.sum(axis=0), "o-", label="Scheduled", color="#1b7837")
    ax.plot(days, [x["objective"] for x in q2], "s--", label="Required", color="#762a83")
    ax.set_xlabel("Day"); ax.set_ylabel("Workers"); ax.grid(alpha=.2)
    ax.legend(frameon=False)
    fig.tight_layout(); fig.savefig(FIG / "monthly_roster.pdf", bbox_inches="tight"); plt.close(fig)


def main():
    q1 = [solve_day(d, False) for d in range(30)]
    q2 = [solve_day(d, True) for d in range(30)]
    requirements = [x["objective"] for x in q2]
    n_workers, roster = solve_monthly_roster(requirements)
    allocation = assign_shifts(q2, roster)
    sensitivity = {}
    for label, factor in (("minus5", 0.95), ("base", 1.0), ("plus5", 1.05)):
        sols = q2 if factor == 1.0 else [solve_day(d, True, 25 * factor, 10 * factor) for d in range(30)]
        sensitivity[label] = {
            "factor": factor,
            "worker_days": sum(x["objective"] for x in sols),
            "peak": max(x["objective"] for x in sols),
        }
    save_outputs(q1, q2, n_workers, roster, allocation, sensitivity)
    print(f"Q1 worker-days={sum(x['objective'] for x in q1)}, max={max(x['objective'] for x in q1)}")
    print(f"Q2 worker-days={sum(x['objective'] for x in q2)}, max={max(x['objective'] for x in q2)}")
    print(f"Q3 minimum workers={n_workers}, all roster checks passed")
    print("Sensitivity:", sensitivity)


if __name__ == "__main__":
    main()
