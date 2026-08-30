from pathlib import Path
import json
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parent
GEN=ROOT/'generated'; FIG=ROOT/'paper'/'figures'; FIG.mkdir(exist_ok=True)
r=json.loads((GEN/'results.json').read_text(encoding='utf-8'))
wb=load_workbook(ROOT/'附件.xlsx',data_only=True); ws=wb.active
a=np.zeros((30,24),dtype=int)
for row in ws.iter_rows(min_row=2,values_only=True): a[int(row[0])-1,int(row[1])]=int(row[2])

def tex_escape(s): return str(s).replace('_',r'\_').replace('%',r'\%')

def write_daily_detail():
    lines=[r'\begin{longtable}{rrrrrr}',r'\caption{逐日班次与人数详细结果}\label{tab:daily-detail}\\',r'\toprule 日 & 问题 & 班次起点 & 各班次人数 & 总人数 & 到货量\\\midrule',r'\endfirsthead\toprule 日 & 问题 & 班次起点 & 各班次人数 & 总人数 & 到货量\\\midrule\endhead']
    for q in ('q1','q2'):
        for x in r[q]:
            starts=x['selected']; nums=[x['workers'][str(s)] for s in starts]
            lines.append(f"{x['day']} & {q.upper()} & {'/'.join(map(str,starts))} & {'/'.join(map(str,nums))} & {x['objective']} & {int(a[x['day']-1].sum())}\\\\")
    lines += [r'\bottomrule',r'\end{longtable}']; (GEN/'daily_detail.tex').write_text('\n'.join(lines),encoding='utf-8')

def write_hourly_detail():
    lines=[r'\begin{longtable}{rrrrrr}',r'\caption{问题二逐小时流量平衡核验}\label{tab:hourly-detail}\\',r'\toprule 日 & 时 & 到货量 & 处理能力 & 实际处理 & 期末库存\\\midrule',r'\endfirsthead\toprule 日 & 时 & 到货量 & 处理能力 & 实际处理 & 期末库存\\\midrule\endhead']
    for x in r['q2']:
        for h in range(24):
            lines.append(f"{x['day']} & {h} & {int(a[x['day']-1,h])} & {x['capacity'][h]:.1f} & {x['processing'][h]:.1f} & {x['backlog'][h]:.1f}\\\\")
    lines += [r'\bottomrule',r'\end{longtable}']; (GEN/'hourly_detail.tex').write_text('\n'.join(lines),encoding='utf-8')

def write_roster_detail():
    ws=load_workbook(GEN/'排班结果.xlsx',data_only=True)['问题三出勤表']
    lines=[r'\begin{longtable}{r*{10}{c}r}',r'\caption{问题三人员出勤矩阵（每页按10天展示）}\label{tab:roster-detail}\\',r'\toprule 工人 & 1--3日 & 4--6日 & 7--9日 & 10--12日 & 13--15日 & 16--18日 & 19--21日 & 22--24日 & 25--27日 & 28--30日 & 工日\\\midrule',r'\endfirsthead\toprule 工人 & 1--3日 & 4--6日 & 7--9日 & 10--12日 & 13--15日 & 16--18日 & 19--21日 & 22--24日 & 25--27日 & 28--30日 & 工日\\\midrule\endhead']
    for i in range(2,ws.max_row+1):
        vals=[int(ws.cell(i,d).value) for d in range(2,32)]
        chunks=[''.join(map(str,vals[j:j+3])) for j in range(0,30,3)]
        lines.append(f"{ws.cell(i,1).value} & "+' & '.join(chunks)+f" & {sum(vals)}\\\\")
    lines += [r'\bottomrule',r'\end{longtable}']; (GEN/'roster_detail.tex').write_text('\n'.join(lines),encoding='utf-8')

def write_lowhour_detail():
    lines=[r'\begin{longtable}{rrrr}',r'\caption{问题二低效率小时分配}\label{tab:lowhour-detail}\\',r'\toprule 日 & 班次起点 & 相对小时 & 人数\\\midrule',r'\endfirsthead\toprule 日 & 班次起点 & 相对小时 & 人数\\\midrule\endhead']
    for x in r['q2']:
        for s,ks in x['low_productivity'].items():
            for k,n in ks.items(): lines.append(f"{x['day']} & {s} & {k} & {n}\\\\")
    lines += [r'\bottomrule',r'\end{longtable}']; (GEN/'lowhour_detail.tex').write_text('\n'.join(lines),encoding='utf-8')

def charts():
    days=np.arange(1,31); q1=np.array([x['objective'] for x in r['q1']]); q2=np.array([x['objective'] for x in r['q2']])
    # capacity utilization by hour across Q2
    cap=np.array([x['capacity'] for x in r['q2']]); proc=np.array([x['processing'] for x in r['q2']]); util=np.divide(proc,cap,out=np.zeros_like(proc),where=cap>0)
    fig,ax=plt.subplots(figsize=(9,4.6)); ax.plot(range(24),util.mean(0)*100,'o-',color='#2166ac'); ax.fill_between(range(24),np.percentile(util,10,0)*100,np.percentile(util,90,0)*100,alpha=.2,color='#67a9cf'); ax.set(xlabel='Hour',ylabel='Utilization (%)',title='Q2 mean hourly capacity utilization and 10-90% band'); ax.set_xticks(range(0,24,2)); ax.grid(alpha=.2); fig.tight_layout(); fig.savefig(FIG/'capacity_utilization.pdf'); plt.close(fig)
    # backlog profiles
    back=np.array([x['backlog'] for x in r['q2']]); fig,ax=plt.subplots(figsize=(9,4.6)); ax.plot(range(24),back.mean(0),color='#b2182b',lw=2,label='Mean backlog'); ax.fill_between(range(24),np.percentile(back,10,0),np.percentile(back,90,0),alpha=.2,color='#ef8a62',label='10-90% band'); ax.axvline(16,color='k',ls='--',lw=1,label='16:00'); ax.set(xlabel='Hour',ylabel='Backlog (items)',title='Q2 hourly backlog profile'); ax.legend(frameon=False); ax.grid(alpha=.2); fig.tight_layout(); fig.savefig(FIG/'backlog_profile.pdf'); plt.close(fig)
    # shift start histogram
    counts={s:0 for s in range(17)}
    for x in r['q2']:
        for s in x['selected']: counts[s]+=1
    fig,ax=plt.subplots(figsize=(9,4.2)); ax.bar(list(counts),list(counts.values()),color='#4d9221'); ax.set(xlabel='Shift start hour',ylabel='Frequency across 30 days',title='Q2 selected shift-start frequency'); ax.set_xticks(range(17)); ax.grid(axis='y',alpha=.2); fig.tight_layout(); fig.savefig(FIG/'shift_start_frequency.pdf'); plt.close(fig)
    # daily incremental cost
    fig,ax=plt.subplots(figsize=(9,4.2)); ax.bar(days,q2-q1,color='#d6604d'); ax.axhline((q2-q1).mean(),color='#762a83',ls='--',label='Mean increment'); ax.set(xlabel='Day',ylabel='Additional workers',title='Additional workers caused by Q2 constraints'); ax.legend(frameon=False); ax.grid(axis='y',alpha=.2); fig.tight_layout(); fig.savefig(FIG/'constraint_increment.pdf'); plt.close(fig)
    # roster heatmap (full roster)
    ws=load_workbook(GEN/'排班结果.xlsx',data_only=True)['问题三出勤表']
    rost=np.array([[int(ws.cell(i,d).value) for d in range(2,32)] for i in range(2,ws.max_row+1)])
    fig,ax=plt.subplots(figsize=(9,9)); ax.imshow(rost,aspect='auto',cmap='Greens',interpolation='nearest'); ax.set(xlabel='Day',ylabel='Worker ID',title='Q3 monthly attendance matrix'); ax.set_xticks(range(0,30,3)); ax.set_xticklabels(range(1,31,3)); fig.tight_layout(); fig.savefig(FIG/'roster_heatmap.pdf'); plt.close(fig)

write_daily_detail(); write_hourly_detail(); write_roster_detail(); write_lowhour_detail(); charts()
print('enhanced tables and five figures generated')
